# This script can automate the following things:
# 1. Replace the bot author info in the release note table with the actual PR authors.
# 2. Add the history duplicated release notes based on issue links and author info. The duplicate release notes in the same series will be not added. For example, if you are working on v6.5.4 release notes, the notes from other v6.5.x with the same issue number will not be counted and added as duplicated notes.
# 3. Make a copy of the patch release note template file and write the duplicated release notes to the copy.

# Before running this script, you need to get a GitHub personal access token (https://docs.github.com/en/authentication/keeping-your-account-and-data-secure/creating-a-personal-access-token) and save it in a text file.

# -*- coding: utf-8 -*-

from github import Github
import re
import openpyxl
import os
import shutil

version = '6.5.3' # Specifies the target TiDB version
release_note_excel = r'/Users/userid/Downloads/download_tirelease_tmp_patch_6.5.3_release_note_2023-06-06.xlsx' # Specifies the path of release note table with PR links and issue links
ext_path = r'/Users/userid/Documents/GitHub/mygithubid/docs-cn/releases'  # Specifies the path of the existing release notes
template_file = r'/Users/userid/Documents/GitHub/mygithubid/docs/resources/doc-templates/patch_release_note_template_zh.md' # Specifies the path of the release note template file

with open("/Users/userid/Documents/PingCAP/Python_scripts/GitHub/gh_token2.txt", "r") as f: # Read the GitHub personal access token from the token.txt file
    access_token = f.read().strip()

# Get the issue info of the existing release notes
def store_exst_rn(ext_path, version):

    exst_notes = []
    exst_issue_nums = []
    exst_note_levels = []
    release_file = os.path.join(ext_path, f'release-{version}.md')

    version_parts = version.split('.')
    major_minor_version = '.'.join(version_parts[:2])

    for maindir, subdir, files in os.walk(ext_path):
        for afile in files:
            file_path = (os.path.join(maindir, afile))
            if file_path.endswith('.md') and major_minor_version not in afile: # Exclude duplicate notes that are in the same major or minor releases. For example, excluding 6.5.x dup release notes for v6.5.3
                with open(file_path,'r', encoding='utf-8') as fp:
                    level1 = level2 = level3 = ""
                    for line in fp:
                        exst_issue_num = re.search(r'https://github.com/(pingcap|tikv)/[\w-]+/(issues|pull)/\d+', line)
                        authors = re.findall(r'@\[([^\]]+)\]', line) # Get the list of authors in this line
                        if exst_issue_num:
                            if exst_issue_num.group() not in exst_issue_nums:
                                note_level = level1 + level2 + level3
                                note_pair = [exst_issue_num.group(),line.strip(),afile, note_level, authors]
                                exst_issue_nums.append(exst_issue_num.group())
                                exst_notes.append(note_pair)
                            else:
                                continue
                        elif line.startswith("##"):
                            level1 = "> " + line.replace("##","").strip()
                            level2 = level3 = ""
                        elif (line.startswith ("+") or line.startswith ("-")) and (not authors):
                            level2 = "> " + line.replace("+","").replace("-","").strip()
                            level3 = ""
                        elif (line.startswith ("    +") or line.startswith ("    -")) and (not authors):
                            level3 = "> " + line.replace("    +","").replace("    -","").strip()
                        else:
                            continue
            else:
                pass

    if len(exst_issue_nums) != 0:
        return exst_notes
    else:
        return 0

def get_pr_info_from_github(cp_pr_link,cp_pr_title, current_pr_author):

    g = Github(access_token, timeout=30)# Create a Github object with the access token
    target_pr_number_existence = 1
    target_repo_pr_link= cp_pr_link.rsplit('/', 1)[0]
    target_pr_number = re.findall(r'\(#(\d+)\)$', cp_pr_title) # Match the original PR number in the end of the cherry-pick PR

    if target_pr_number: # Match the original PR number according to the title of the cherry-pick PR
        pass
    else: # Match the original PR number according to the head branch name of the cherry-pick PR
        cp_pr_info = cp_pr_link.split("/")
        owner, repo, cp_pr_number = cp_pr_info[-4], cp_pr_info[-3], cp_pr_info[-1]
        repo_obj = g.get_repo(f"{owner}/{repo}")
        try:
            cp_pr_obj = repo_obj.get_pull(int(cp_pr_number))
            cp_head_branch = cp_pr_obj.head.ref
            target_pr_number = re.findall(r'cherry-pick-(\d+)', cp_head_branch)
            if target_pr_number:
                pass
            else:
                target_pr_number_existence = 0
        except:
            target_pr_number_existence = 0

    if target_pr_number_existence == 1:
        target_pr_link = target_repo_pr_link + '/' + target_pr_number[0]
        pr_info = target_pr_link.split("/")
        owner, repo, target_pr_number = pr_info[-4], pr_info[-3], pr_info[-1]
        repo_obj = g.get_repo(f"{owner}/{repo}")# Get the repository object
        try:
            pr_obj = repo_obj.get_pull(int(target_pr_number))# Get the pull request object
            pr_author = pr_obj.user.login # Get the author of the pull request
        except:
            print("Failed to get the original PR information for this PR: " + cp_pr_link)
    else:
        pr_author = current_pr_author # Use the current author if the cherry-pick PR cannot be found

    return(pr_author)

def update_pr_author_and_release_notes(excel_path):

    # Open the excel file
    workbook = openpyxl.load_workbook(excel_path)
    # Specify the target sheet
    sheet = workbook['pr_for_release_note']

    # Get the sheet header
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)) # Read the first line in the sheet

    # Get the column info
    pr_author_index = header.index('pr_author')
    pr_link_index = header.index('pr_link')
    pr_title_index = header.index('pr_title')
    pr_formated_rn_index = header.index('formated_release_note')
    pr_last_col_index = sheet.max_column
    sheet.insert_cols(pr_last_col_index + 1) # Insert a new column for the dup release notes
    sheet.cell(row=1, column=pr_last_col_index + 1, value='published_release_notes') # Set a column name
    # Go through each row
    dup_notes = []
    dup_notes_levels = []
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        # If pr_author is ti-chi-bot or ti-srebot
        current_pr_author = row[pr_author_index]
        current_formated_rn= row[pr_formated_rn_index]
        if current_pr_author in ['ti-chi-bot', 'ti-srebot']:
           print ("Replacing the author info for row " + str(row_index) + ".")
           actual_pr_author = get_pr_info_from_github(row[pr_link_index], row[pr_title_index], current_pr_author) # Get the PR author according to the cherry-pick PR
           pr_author_cell = sheet.cell(row=row_index, column=pr_author_index+1, value = actual_pr_author)#Fill in the pr_author_cell
           updated_formated_rn = current_formated_rn.replace("[{}](https://github.com/{}".format(current_pr_author, current_pr_author),"[{}](https://github.com/{}".format(actual_pr_author, actual_pr_author))
           formated_release_note_cell = sheet.cell(row=row_index, column=pr_formated_rn_index+1, value = updated_formated_rn) # Fill in the formated_release_note_cell
           current_pr_author = actual_pr_author
        else:
            pass

        ## Add the dup release note info
        issue_link = re.search('https://github.com/(pingcap|tikv)/[\w-]+/issues/\d+', current_formated_rn)
        if issue_link:
            for note_pair in note_pairs:
                if (issue_link.group() == note_pair[0]) and ((current_pr_author in note_pair[4]) or len(note_pair[4]) == 0): # Add the dup release notes only if the issues link is the same as the existing one and the current author is in the existing author list
                    #print('A duplicated note is found in row ' + str(row_index) + " from " + note_pair[2] + note_pair[1])
                    dup_formated_rn = '- (dup): {} {} {}'.format(note_pair[2], note_pair[3], note_pair[1])
                    #print (note_pair)
                    sheet.cell(row=row_index, column=pr_last_col_index+1, value=dup_formated_rn)
                    if dup_formated_rn not in dup_notes: # Collect the dup release note if it is not collected before
                        dup_notes.append(dup_formated_rn)
                        print ("-----")
                        print (dup_formated_rn)
                        dup_notes_level = note_pair[3].replace("Bug 修复", "错误修复")
                        dup_notes_levels.append(dup_notes_level)
                    else:
                        pass
                else:
                    pass
        elif (not issue_link) and ("/issue/" in current_formated_rn):
            print(current_formated_rn)
        else:
            pass

    workbook.save(release_note_excel)
    return dup_notes, dup_notes_levels

# Add the dup release notes to the release note file
def create_release_file(version, dup_notes_levels, dup_notes):

    release_file = os.path.join(ext_path, f'release-{version}.md')
    shutil.copyfile(template_file, release_file)
    # Replace the file content
    with open(release_file, 'r+') as file:
        content = file.read()
        content = content.replace('x.y.z', version)
        version_parts = version.split('.')
        major_minor_version = '.'.join(version_parts[:2])
        content = content.replace('x.y', major_minor_version)
        level1 = level2 = level3 = ""
        lines = content.splitlines()
        newlines = []
        note_level = ""
        note_levels = []
        other_dup_notes = []
        for line in lines:
            if "placeholder" in line:
                note_level = level1 + level2 + level3
                note_levels.append(note_level)
                newline = line.replace("- placeholder", "")
                for dup_note_level, dup_note in zip(dup_notes_levels, dup_notes): # Add the dup release notes to the release note file
                    if dup_note_level == note_level:
                        newlines.append(newline+dup_note)
                    else:
                        continue
                if "Other dup notes" in note_level: # Add the dup release notes without corresponding categories to the release note file
                    for dup_note_level, dup_note in zip(dup_notes_levels, dup_notes):
                        if dup_note_level not in note_levels:
                            newlines.append(newline+dup_note)
                            other_dup_notes.append(dup_note)
                    if len(other_dup_notes) == 0:
                        newlines = newlines[:-2] # Remove the last two lines if other dup notes do not exist
                    else:
                        pass
            elif line.startswith("##"):
                level1 = "> " + line.replace("##","").strip()
                level2 = level3 = ""
                newlines.append(line)
            elif (line.startswith ("+") or line.startswith ("-")) and "GitHub ID" not in line:
                level2 = "> " + line.replace("+","").replace("-","").strip()
                level3 = ""
                newlines.append(line)
            elif (line.startswith ("    +") or line.startswith ("    -")) and "GitHub ID" not in line:
                level3 = "> " + line.replace("    +","").replace("    -","").strip()
                newlines.append(line)
            else:
                newlines.append(line)

        #print(note_levels)
        content = "\n".join(newlines)
        file.seek(0)
        file.write(content)
        file.truncate()
        print(f'The v{version} release note is now created in the following directory: \n {release_file}')

if __name__ == '__main__':
    note_pairs = store_exst_rn(ext_path, version)
    dup_notes, dup_notes_levels = update_pr_author_and_release_notes(release_note_excel)
    print ("The bot author info in the excel is now replaced with the actual authors.")
    version_parts = version.split('.')
    if len(version_parts) >= 2:
        create_release_file(version, dup_notes_levels, dup_notes)